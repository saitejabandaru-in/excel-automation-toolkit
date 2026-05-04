from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .data_loader import load_table
from .feature_engineering import add_time_features, aggregate_metrics
from .preprocessing import clean_dataframe, coerce_numeric_columns, normalize_column_name


def build_report(
    input_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str | int | None = 0,
    required_columns: Iterable[str] = (),
    date_columns: Iterable[str] = (),
    numeric_columns: Iterable[str] = (),
    dimensions: Iterable[str] = (),
    amount_column: str | None = None,
) -> Path:
    """Create a formatted Excel report with raw, cleaned, summary, and metadata sheets."""
    raw = load_table(input_path, sheet_name=sheet_name)
    cleaned = clean_dataframe(raw, required_columns=required_columns, date_columns=date_columns)

    numeric_inputs = list(numeric_columns)
    if amount_column:
        numeric_inputs.append(amount_column)
    cleaned = coerce_numeric_columns(cleaned, numeric_inputs)
    enriched = add_time_features(cleaned, date_columns)

    summary = _build_summary(enriched, dimensions=dimensions, amount_column=amount_column)
    metadata = pd.DataFrame(
        [
            {"key": "input_path", "value": str(input_path)},
            {"key": "rows_loaded", "value": len(raw)},
            {"key": "rows_processed", "value": len(enriched)},
            {"key": "generated_at_utc", "value": datetime.now(timezone.utc).isoformat(timespec="seconds")},
        ]
    )

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        raw.to_excel(writer, sheet_name="Raw Data", index=False)
        enriched.to_excel(writer, sheet_name="Processed Data", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        metadata.to_excel(writer, sheet_name="Metadata", index=False)

    _format_workbook(destination)
    return destination


def _build_summary(
    df: pd.DataFrame,
    *,
    dimensions: Iterable[str],
    amount_column: str | None,
) -> pd.DataFrame:
    normalized_dimensions = [normalize_column_name(column) for column in dimensions if column]
    normalized_amount = normalize_column_name(amount_column) if amount_column else None

    if normalized_dimensions and normalized_amount and normalized_amount in df.columns:
        return aggregate_metrics(
            df,
            dimensions=normalized_dimensions,
            metrics={normalized_amount: ["count", "sum", "mean"]},
        )

    return pd.DataFrame(
        [
            {"metric": "rows", "value": len(df)},
            {"metric": "columns", "value": len(df.columns)},
        ]
    )


def _format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1C2B36")
    header_font = Font(color="E6EEF3", bold=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            continue

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)

        table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        table_name = "".join(part for part in worksheet.title.title() if part.isalnum())[:240] or "ReportTable"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        worksheet.freeze_panes = "A2"

    workbook.save(path)

