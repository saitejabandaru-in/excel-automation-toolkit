from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple


def sync_dataframe_to_workbook(
    workbook_path: str | Path,
    df: pd.DataFrame,
    *,
    sheet_name: str = "Processed Data",
    start_cell: str = "A1",
    clear_existing: bool = True,
) -> Path:
    """Write a dataframe into an Excel workbook sheet for dashboard consumption."""
    path = Path(workbook_path)
    workbook = load_workbook(path) if path.exists() else Workbook()

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if clear_existing:
            worksheet.delete_rows(1, worksheet.max_row)
    else:
        worksheet = workbook.create_sheet(sheet_name)

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1 and workbook["Sheet"].max_row == 1:
        del workbook["Sheet"]

    start_row, start_column = coordinate_to_tuple(start_cell)
    for offset, column in enumerate(df.columns):
        worksheet.cell(row=start_row, column=start_column + offset, value=column)

    for row_index, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for column_index, value in enumerate(row, start=start_column):
            worksheet.cell(row=row_index, column=column_index, value=value)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path

