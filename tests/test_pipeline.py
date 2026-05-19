from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from excel_automation_toolkit.cli import main
from excel_automation_toolkit.data_loader import load_table
from excel_automation_toolkit.pipeline import PipelineConfig, run_pipeline
from excel_automation_toolkit.preprocessing import clean_dataframe


def test_clean_dataframe_normalizes_headers_and_trims_text() -> None:
    df = pd.DataFrame(
        {
            " Order Date ": ["2026-01-01"],
            " Region ": [" North "],
            "Revenue": ["1,250"],
        }
    )

    cleaned = clean_dataframe(df, required_columns=["order_date"], date_columns=["order_date"])

    assert list(cleaned.columns) == ["order_date", "region", "revenue"]
    assert cleaned.loc[0, "region"] == "North"
    assert pd.Timestamp("2026-01-01") == cleaned.loc[0, "order_date"]


def test_pipeline_generates_formatted_workbook(tmp_path: Path) -> None:
    input_path = tmp_path / "sales.csv"
    output_path = tmp_path / "report.xlsx"
    input_path.write_text(
        "Order Date,Region,Revenue\n"
        "2026-01-01,North,100\n"
        "2026-01-02,North,250\n"
        "2026-01-03,South,75\n",
        encoding="utf-8",
    )

    result = run_pipeline(
        PipelineConfig(
            input_path=input_path,
            output_path=output_path,
            date_columns=("Order Date",),
            numeric_columns=("Revenue",),
            dimensions=("Region",),
            amount_column="Revenue",
        )
    )

    assert result == output_path
    workbook = load_workbook(output_path, read_only=False)
    assert {"Raw Data", "Processed Data", "Summary", "Metadata"} <= set(workbook.sheetnames)

    summary = workbook["Summary"]
    rows = list(summary.iter_rows(values_only=True))
    assert rows[0] == ("region", "revenue_count", "revenue_sum", "revenue_mean")
    assert ("North", 2, 350, 175) in rows


def test_cli_run_generates_report(tmp_path: Path, capsys) -> None:
    input_path = tmp_path / "sales.csv"
    output_path = tmp_path / "cli-report.xlsx"
    input_path.write_text("Region,Revenue\nNorth,100\nSouth,75\n", encoding="utf-8")

    exit_code = main(
        [
            "run",
            "--input",
            str(input_path),
            "--output",
            str(output_path),
            "--numeric-column",
            "Revenue",
            "--dimension",
            "Region",
            "--amount-column",
            "Revenue",
        ]
    )

    assert exit_code == 0
    assert output_path.exists()
    assert f"Report generated: {output_path}" in capsys.readouterr().out


def test_loading_all_excel_sheets_raises_clear_error(tmp_path: Path) -> None:
    input_path = tmp_path / "multi-sheet.xlsx"
    with pd.ExcelWriter(input_path, engine="openpyxl") as writer:
        pd.DataFrame({"Region": ["North"]}).to_excel(writer, sheet_name="Sales", index=False)

    try:
        load_table(input_path, sheet_name=None)
    except ValueError as exc:
        assert "Expected a single worksheet" in str(exc)
        assert "Sales" in str(exc)
    else:
        raise AssertionError("Expected load_table to reject multiple worksheet output")
